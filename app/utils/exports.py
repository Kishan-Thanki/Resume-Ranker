import io
import csv
import openpyxl
from io import BytesIO
from typing import List, Dict
from openpyxl.utils import get_column_letter

def generate_csv_ranked_resumes(ranked_resumes_list: List[Dict]) -> bytes:
    output = io.StringIO()
    fieldnames = [
        "uuid", "filename", "combined_score", "skill_score",
        "text_score", "experience_score", "skills_found", "experience_years",
        "contact_info"
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for row in ranked_resumes_list:
        writer.writerow(row)

    return output.getvalue().encode("utf-8")

def generate_excel_from_ranked_data(ranked_resumes_list: List[Dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Resumes"

    headers = [
        "uuid", "filename", "combined_score", "skill_score",
        "text_score", "experience_score", "skills_found", "experience_years",
        "contact_info"
    ]
    ws.append(headers)

    for row in ranked_resumes_list:
        contact_info = row.get("contact_info", {})
        contact_string = f"Email: {contact_info.get('email', '')}, Phone: {contact_info.get('phone', '')}, Location: {contact_info.get('location', '')}"

        ws.append([
            row.get("uuid", ""),
            row.get("filename", ""),
            row.get("combined_score", 0),
            row.get("skill_score", 0),
            row.get("text_score", 0),
            row.get("experience_score", 0),
            row.get("skills_found", ""),
            row.get("experience_years", 0),
            contact_string
        ])

    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].auto_size = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()